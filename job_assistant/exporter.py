from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import Job
from .urls import canonicalize_job_url

HEADERS = [
    "Date", "Platform", "Company", "Role", "URL", "Location", "Salary", "Match Score",
    "Skills Missing for 100% Match", "Status", "Notes", "Follow-up Date", "Followed Up",
]


def export_jobs(jobs: list[Job], output: Path) -> Path:
    """Create a readable, de-duplicated application tracker workbook."""
    output.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    sheet.title = "Job Tracker"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:M1")
    title = sheet["A1"]
    title.value = "Linkedin Job Application Assistant"
    title.font = Font(size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="1F4E78")
    title.alignment = Alignment(horizontal="left")
    sheet.row_dimensions[1].height = 28
    sheet.append([])
    sheet.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="2F75B5")
    for cell in sheet[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    seen_urls: set[str] = set()
    for job in jobs:
        canonical_url = canonicalize_job_url(job.url)
        if canonical_url in seen_urls:
            continue
        seen_urls.add(canonical_url)
        sheet.append([
            _as_date(job.date_found), job.platform or _platform_from_url(job.url), job.company, job.title,
            canonical_url, job.location, job.salary, job.match_score, job.missing_skills, job.status, job.notes,
            _as_date(job.follow_up_date), "Yes" if job.followed_up else "No",
        ])
    final_row = max(sheet.max_row, 3)
    if final_row > 3:
        table = Table(displayName="JobApplications", ref=f"A3:M{final_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
        for row in range(4, final_row + 1):
            url_cell = sheet.cell(row, 5)
            url_cell.hyperlink = url_cell.value
            url_cell.style = "Hyperlink"
            sheet.cell(row, 1).number_format = "yyyy-mm-dd"
            sheet.cell(row, 12).number_format = "yyyy-mm-dd"
            sheet.cell(row, 9).alignment = Alignment(vertical="top", wrap_text=True)
            sheet.cell(row, 11).alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A4"
    widths = [13, 14, 24, 28, 48, 22, 16, 13, 34, 25, 64, 16, 14]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.column_dimensions["K"].width = 42
    # Excel tables already own their filter range. A second worksheet-level
    # filter over the same cells makes Excel repair the generated workbook.
    if final_row > 3:
        score_range = f"H4:H{final_row}"
        sheet.conditional_formatting.add(score_range, CellIsRule(operator="greaterThanOrEqual", formula=["75"], fill=PatternFill("solid", fgColor="C6E0B4")))
        sheet.conditional_formatting.add(score_range, CellIsRule(operator="lessThan", formula=["35"], fill=PatternFill("solid", fgColor="F4CCCC")))
        status_validation = DataValidation(type="list", formula1='"saved,reviewing,ready_for_manual_submit,submitted_manually,rejected,archived"', allow_blank=True)
        sheet.add_data_validation(status_validation)
        status_validation.add(f"J4:J{max(final_row, 104)}")
        follow_up_validation = DataValidation(type="list", formula1='"No,Yes"', allow_blank=False)
        sheet.add_data_validation(follow_up_validation)
        follow_up_validation.add(f"M4:M{max(final_row, 104)}")
    book.save(output)
    return output


def _as_date(value: str | None) -> date | str | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return value


def _platform_from_url(url: str) -> str:
    from urllib.parse import urlparse
    hostname = (urlparse(url).hostname or "Unknown").removeprefix("www.")
    return {"linkedin.com": "LinkedIn", "indeed.com": "Indeed", "glassdoor.com": "Glassdoor"}.get(hostname, hostname.title())
